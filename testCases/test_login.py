import base64
import sys
import time

import openpyxl
import pytest
import unittest
from selenium import webdriver
from pages.LoginPage import Login
from utilities.customLogger import LogGen


class Test_001_Login:
    # username = "zaloom7@impaqtive.com"
    # pas = base64.b64decode(b'emFsb29tQDc3').decode('utf-8')
    logger = LogGen.loggen()

    # def screen_shot(self, setup):
    #     """Take a Screen-shot of the drive homepage, when it Failed."""
    #     self.driver=setup
    #     for method, error in self._outcome.errors:
    #         if error:
    #             self.driver.get_screenshot_as_file("Screenshots" + self.id() + ".png")

    def test_login(self, setup):

        self.logger.info("*************** Test_001_Login *****************")
        self.logger.info("****Started login test ****")
        self.driver = setup
        self.logger.info("****Opening URL****")
        self.lp = Login(self.driver)

        path = "C:\\selenium browser drivers\\login.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name("Sheet1")
        rows = sheet.max_row
        # cols = sheet.max_coloumn

        for r in range(2, rows + 1):
            self.driver.get("https://login.salesforce.com")

            # print(sheet.cell(r, 1).value)
            username = sheet.cell(r, 1).value
            password = sheet.cell(r, 2).value
            self.lp.setUsername(username)
            time.sleep(2)
            self.lp.setPassword(password)
            self.lp.clickLogin()
            x = self.driver.title
            if x == "Lightning Experience":
                self.logger.info("**** Home page title test passed ****")
                time.sleep(6)
                assert True
            else:
                self.logger.error("**** Home page title test failed****")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_login.png")
                time.sleep(2)
                assert False
        self.logger.info("**************** Completed  test_login ************* ")

# if __name__ == '__main__':
#     unittest.main(verbosity=2)
