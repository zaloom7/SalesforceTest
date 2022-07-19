import base64
import time

import openpyxl
import pytest
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pages.LoginPage import Login
from pages.AddRecord import Record
from utilities.customLogger import LogGen


class Test_003:
    username = "zaloom7@impaqtive.com"

    # pas = base64.b64decode(b'emFsb29tQDc3').decode('utf-8')

    pas = base64.b64decode(b'emFsb29tQDc3Nw==').decode('utf-8')
    logger = LogGen.loggen()

    def test_record(self, setup):
        self.logger.info("*************** Test_001_Login *****************")
        self.logger.info("****Started Home page title test ****")
        self.driver = setup
        self.logger.info("****Opening URL****")
        self.driver.get("https://login.salesforce.com")
        self.lp = Login(self.driver)
        self.lp.setUsername(self.username)
        time.sleep(1)
        self.lp.setPassword(self.pas)
        print(self.pas)
        time.sleep(1)
        self.lp.clickLogin()
        self.addrec = Record(self.driver)
        time.sleep(5)
        self.driver.implicitly_wait(12)
        self.addrec.clickAppLauncher()
        time.sleep(2)
        # self.addrec.serviceConsole()
        # time.sleep(2)
        # self.addrec.dropdown()
        # time.sleep(2)
        # addrec.clickContacts()
        # addrec.clickNew()

        path = "C:\\selenium browser drivers\\Contacts.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb["Sheet1"]
        rows = sheet.max_row
        cols = sheet.max_column
        search = sheet.cell(1, 1).value
        time.sleep(4)
        self.addrec.searchTab(search)
        self.addrec.pinAllRecords(search)
        for r in range(4, rows + 1):
            # print(sheet.cell(j, 1).value)
            lastname = sheet.cell(r, 1).value
            time.sleep(1)

            try:
                self.addrec.searchName(lastname)
                self.addrec.editDetails(lastname)
            except NoSuchElementException:
                time.sleep(2)
                self.addrec.clickNew()
                time.sleep(2)
            savename = str(r - 3) + '.png'
            # addrec.newContact()
            # time.sleep(1)
            for c in range(1, cols):
                y = sheet.cell(2, c).value
                t = sheet.cell(3, c).value
                value = sheet.cell(r, c).value
                try:
                    if t == "text":
                        self.addrec.enterValues(y, value)
                    elif t == "Lookup":
                        self.addrec.selectLookup(y, value)
                    elif t == "Picklist":
                        self.addrec.selectPicklist(y, value)
                except NoSuchElementException:
                    pass
            self.addrec.save()
            time.sleep(2)
            # self.addrec.leadSource()
            # time.sleep(1)
            # self.addrec.selectPicklist()
            # time.sleep(1)
            # self.addrec.email(email)
            # time.sleep(1)
            # self.addrec.save()
            # time.sleep(2)

            try:
                self.driver.implicitly_wait(5)
                self.addrec.error()
                time.sleep(3)
                sheet.cell(r, cols).value = "Fail"
                self.driver.save_screenshot(".\\Screenshots\\" + savename)
            except NoSuchElementException:
                sheet.cell(r, cols).value = "Pass"

                # try:
                #     self.driver.implicitly_wait(5)
                #     self.addrec.duplicate()
                #     # self.addrec.viewDuplicates()
                #     time.sleep(3)
                #     sheet.cell(r+2, cols).value = "duplicate record"
                #     self.driver.save_screenshot(".\\Screenshots\\" + savename)
                #
                #
                # except NoSuchElementException:
                #     sheet.cell(r+2, cols).value = "Pass"

            wb.save('TestResult.xlsx')
            wb.close()
            try:
                self.addrec.closeContact()
                time.sleep(2)
            except NoSuchElementException:
                pass
            time.sleep(2)
            self.driver.delete_all_cookies()
            self.driver.execute_script('arguments[0].click();',
                                       self.driver.find_element(by=By.XPATH, value="//a[@title='{}']".format(search)))
            time.sleep(3)
        self.driver.close()
        # addrec.clickContacts()
        # time.sleep(1)
        # addrec.yes()
