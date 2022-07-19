import time
import unittest

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from pages.LoginPage import Login


class TestDemo(unittest.TestCase):

    def setUp(self):
        """Set up the chrome Browser and the Tear Down."""
        s = Service(executable_path="C:\\selenium browser drivers\\chromedriver.exe")
        self.driver = webdriver.Chrome(service=s)
        self.driver.delete_all_cookies()
        # NOTE: In addCleanup, the first in, is executed last.
        self.addCleanup(self.screen_shot)
        self.driver.implicitly_wait(5)

    def screen_shot(self):
        """Take a Screen-shot of the drive homepage, when it Failed."""
        for method, error in self._outcome.errors:
            if error:
                self.driver.get_screenshot_as_file(".\\Screenshots\\"+"Screenshots" + self.id() + ".png")

    def test_login(self):
        """A test case that fails because of missing element."""
        self.lp = Login(self.driver)
        path = "C:\\selenium browser drivers\\login.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name("Sheet1")
        rows = sheet.max_row
        # cols = sheet.max_coloumn

        for r in range(2, rows + 1):
            self.driver.get("https://login.salesforce.com")

            # print(sheet.cell(r, 1).value)
            username = sheet.cell(r, 1).value
            password = sheet.cell(r, 2).value
            self.lp.setUsername(username)
            time.sleep(2)
            self.lp.setPassword(password)
            self.lp.clickLogin()
            x = self.driver.title
            if x == "Lightning Experience":
                # self.logger.info("**** Home page title test passed ****")
                time.sleep(6)
                assert True
            else:
                # self.logger.error("**** Home page title test failed****")
                # self.driver.save_screenshot(".\\Screenshots\\" + "test_login.png")
                time.sleep(2)
                assert False


# if __name__ == '__main__':
#     unittest.main(verbosity=2)
