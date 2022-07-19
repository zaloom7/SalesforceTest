import base64
import time

import openpyxl
import pytest
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By

from pages.LoginPage import Login
from pages.AddRecord import Record
from utilities.customLogger import LogGen


class Test_003:
    username = "zaloom7@impaqtive.com"
    # pas = base64.b64decode(b'emFsb29tQDc3').decode('utf-8')

    pas = base64.b64decode(b'emFsb29tQDc3Nw==').decode('utf-8')
    logger = LogGen.loggen()

    def test_record(self, setup):
        self.logger.info("*************** Test_001_Login *****************")
        self.logger.info("****Started Home page title test ****")
        self.driver = setup
        self.logger.info("****Opening URL****")
        self.driver.get("https://login.salesforce.com")
        self.lp = Login(self.driver)
        self.lp.setUsername(self.username)
        time.sleep(1)
        self.lp.setPassword(self.pas)
        print(self.pas)
        time.sleep(1)
        self.lp.clickLogin()
        time.sleep(2)
        self.addrec = Record(self.driver)
        self.driver.implicitly_wait(12)
        self.addrec.clickAppLauncher()
        time.sleep(2)
        self.addrec.serviceConsole()
        time.sleep(2)
        self.addrec.dropdown()
        time.sleep(2)
        # addrec.clickContacts()
        # addrec.clickNew()

        path = "C:\\selenium browser drivers\\Contacts.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb["Sheet1"]
        rows = sheet.max_row

        # cols = sheet.max_coloumn

        for r in range(3, rows + 1):
            # print(sheet.cell(j, 1).value)
            lastname = sheet.cell(r, 1).value
            accName = sheet.cell(r, 2).value
            mobile = sheet.cell(r, 3).value
            email = sheet.cell(r, 4).value
            savename = str(r - 1) + '.png'

            self.addrec.clickContacts()
            time.sleep(2)
            self.driver.find_element(by=By.CSS_SELECTOR, value= ".triggerLinkText").click()
            self.driver.find_element(by=By.XPATH,
                                     value="//span[contains(.,'All Contacts')]").click()
            try:
                self.driver.implicitly_wait(5)
                self.driver.find_element(by=By.XPATH,
                                         value="// a[contains(text(), '{}')]".format(lastname)).click()
                time.sleep(2)
                self.driver.find_element(by=By.XPATH,
                                         value="//a[@data-label='Details']").click()
                time.sleep(1)
                self.driver.find_element(by=By.XPATH,
                                         value="//button[@title='Edit Name']").click()
            except NoSuchElementException:
                self.addrec.clickNew()
                time.sleep(2)
            # addrec.newContact()
            # time.sleep(1)

            try:
                self.addrec.lastName(lastname)
                time.sleep(1)
                self.addrec.searchAccount()
                time.sleep(2)
                self.addrec.accountName(accName)
                time.sleep(1)
                self.driver.implicitly_wait(3)
                # self.addrec.enterButton()
                # time.sleep(1)
                self.addrec.accountclick()
                time.sleep(2)
            except NoSuchElementException:
                pass

                # self.addrec.leadSource()
                # time.sleep(1)
                # self.addrec.selectPicklist()
                # time.sleep(1)
                # self.addrec.email(email)
                # time.sleep(1)
                # self.addrec.save()
                # time.sleep(2)
            self.addrec.leadSource()
            time.sleep(1)
            self.addrec.selectPicklist()
            time.sleep(2)
            self.addrec.email(email)
            time.sleep(2)
            self.addrec.save()
            time.sleep(2)
            try:
                self.driver.implicitly_wait(5)
                self.addrec.error()
                time.sleep(3)
                sheet.cell(r, 5).value = "Fail"
                self.driver.save_screenshot(".\\Screenshots\\" + savename)
            except NoSuchElementException:
                try:
                    self.driver.implicitly_wait(5)
                    self.addrec.duplicate()
                    # self.addrec.viewDuplicates()
                    time.sleep(3)
                    sheet.cell(r, 5).value = "duplicate record"
                    self.driver.save_screenshot(".\\Screenshots\\" + savename)


                except NoSuchElementException:
                    sheet.cell(r, 5).value = "Pass"
            # x = self.addrec.error().text
            # # y = self.addrec.duplicate()
            # if len(x) > 0:
            #     sheet.cell(r, 5).value = "Fail"
            #     self.driver.save_screenshot(".\\Screenshots\\" + savename)
            #
            # else:
            #     self.driver.implicitly_wait(10)
            #     if len(y) > 0:
            #         self.addrec.save()
            #         sheet.cell(r, 5).value = "Duplicates"
            #         self.driver.save_screenshot(".\\Screenshots\\" + savename)
            #     else:
            #         sheet.cell(r, 5).value = "Pass"

            wb.save('TestResult.xlsx')
            wb.close()
            try:
                self.addrec.closeContact()
                time.sleep(2)
                self.driver.implicitly_wait(20)

                self.addrec.discardChanges()
                time.sleep(2)
            except NoSuchElementException:
                pass
        self.driver.close()
        # addrec.clickContacts()
        # time.sleep(1)
        # addrec.yes()

